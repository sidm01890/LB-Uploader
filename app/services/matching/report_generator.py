"""
Report Generator - Process 2
============================
Generates comprehensive Excel and HTML reports for matching results:
1. Executive Summary (match rates, financial summary, exception counts)
2. Detailed Match Report (all matched transactions with full chain)
3. Exception Report (all unmatched/problematic transactions)
4. Store-wise Summary (performance by store)
5. Trend Analysis (daily/weekly match trends)

Created: November 12, 2025
"""

import sys
import os
import logging
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from app.logging_config import setup_logging

setup_logging('INFO')
logger = logging.getLogger('report_generator')


class ReportGenerator:
    """Generates Excel and HTML reports for matching process"""
    
    def __init__(self):
        """Initialize report generator"""
        self.reports_dir = None
    
    def ensure_reports_directory(self, base_path: str = None) -> str:
        """Create reports directory if it doesn't exist"""
        if base_path:
            self.reports_dir = os.path.join(base_path, 'reports', 'matching')
        else:
            # Use project root/reports/matching
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(
                os.path.dirname(os.path.abspath(__file__))
            )))
            self.reports_dir = os.path.join(project_root, 'reports', 'matching')
        
        os.makedirs(self.reports_dir, exist_ok=True)
        logger.info(f"📁 Reports directory: {self.reports_dir}")
        return self.reports_dir
    
    def generate_excel_report(self, 
                            match_summary: Dict, 
                            exception_summary: Dict,
                            matched_records: List[Dict] = None,
                            exception_records: List[Dict] = None) -> str:
        """
        Generate comprehensive Excel report
        
        Args:
            match_summary: Summary from MatchingOrchestrator
            exception_summary: Summary from ExceptionHandler
            matched_records: List of matched transaction records
            exception_records: List of exception records
        
        Returns:
            Path to generated Excel file
        """
        logger.info("📊 Generating Excel report...")
        
        # Ensure reports directory exists
        self.ensure_reports_directory()
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        batch_id = match_summary.get('batch_id', timestamp)
        filename = f"matching_report_{batch_id}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Executive Summary Sheet
        self._create_executive_summary_sheet(wb, match_summary, exception_summary)
        
        # 2. Match Details Sheet
        if matched_records:
            self._create_match_details_sheet(wb, matched_records)
        
        # 3. Exception Details Sheet
        if exception_records:
            self._create_exception_details_sheet(wb, exception_records)
        
        # 4. Store-wise Summary Sheet
        self._create_store_summary_sheet(wb, match_summary, exception_summary)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"  ✅ Excel report saved: {filepath}")
        
        return filepath
    
    def _create_executive_summary_sheet(self, wb: Workbook, match_summary: Dict, exception_summary: Dict):
        """Create executive summary worksheet"""
        ws = wb.create_sheet("Executive Summary")
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_font = Font(bold=True, size=12)
        section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "Matching & Reconciliation Report - Executive Summary"
        ws[f'A{row}'].font = Font(bold=True, size=16)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Report Info
        ws[f'A{row}'] = "Batch ID:"
        ws[f'B{row}'] = match_summary.get('batch_id', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Period:"
        date_range = match_summary.get('date_range', {})
        ws[f'B{row}'] = f"{date_range.get('start')} to {date_range.get('end')}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Generated:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        # Source Data Summary
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "Source Data Summary"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        row += 1
        
        source_counts = match_summary.get('source_counts', {})
        for source, count in source_counts.items():
            ws[f'A{row}'] = source.replace('_', ' ').title()
            ws[f'B{row}'] = count
            row += 1
        row += 1
        
        # Match Results
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "Match Results"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        row += 1
        
        matches = match_summary.get('matches', {})
        for match_type, count in matches.items():
            ws[f'A{row}'] = match_type.replace('_', ' ').title()
            ws[f'B{row}'] = count
            row += 1
        row += 1
        
        # Financial Summary
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "Financial Summary"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        row += 1
        
        financial = match_summary.get('financial', {})
        ws[f'A{row}'] = "Total POS Amount"
        ws[f'B{row}'] = f"₹{financial.get('total_pos_amount', 0):,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Match Rate"
        ws[f'B{row}'] = f"{financial.get('match_rate_percentage', 0):.2f}%"
        ws[f'B{row}'].font = Font(bold=True, color="00B050")
        row += 2
        
        # Exception Summary
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "Exception Summary"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        row += 1
        
        ws[f'A{row}'] = "Total Exceptions"
        ws[f'B{row}'] = exception_summary.get('total_exceptions', 0)
        ws[f'B{row}'].font = Font(bold=True, color="C00000")
        row += 1
        
        ws[f'A{row}'] = "Amount at Risk"
        ws[f'B{row}'] = f"₹{exception_summary.get('total_amount_at_risk', 0):,.2f}"
        ws[f'B{row}'].font = Font(bold=True, color="C00000")
        row += 2
        
        # Exception by Type
        ws[f'A{row}'] = "By Type:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for exc_type, count in exception_summary.get('by_type', {}).items():
            if count > 0:
                ws[f'A{row}'] = exc_type.replace('_', ' ').title()
                ws[f'B{row}'] = count
                row += 1
        row += 1
        
        # Exception by Severity
        ws[f'A{row}'] = "By Severity:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for severity, count in exception_summary.get('by_severity', {}).items():
            if count > 0:
                ws[f'A{row}'] = severity
                ws[f'B{row}'] = count
                
                # Color code by severity
                if severity == 'CRITICAL':
                    ws[f'B{row}'].font = Font(color="C00000", bold=True)
                elif severity == 'HIGH':
                    ws[f'B{row}'].font = Font(color="FF6600")
                elif severity == 'MEDIUM':
                    ws[f'B{row}'].font = Font(color="FFA500")
                
                row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_match_details_sheet(self, wb: Workbook, matched_records: List[Dict]):
        """Create detailed match records worksheet"""
        ws = wb.create_sheet("Match Details")
        
        # Convert to DataFrame
        if matched_records:
            df = pd.DataFrame(matched_records)
            
            # Select and rename columns for readability
            columns_to_show = [
                'recon_date', 'store_name', 'match_type', 'match_status',
                'pos_txn_id', 'pos_amount', 'trm_txn_id', 'trm_amount', 'trm_rrn',
                'mpr_txn_id', 'mpr_amount', 'bank_amount', 'bank_ref',
                'total_variance', 'match_confidence', 'notes'
            ]
            
            # Filter to only existing columns
            columns_to_show = [col for col in columns_to_show if col in df.columns]
            df_display = df[columns_to_show]
            
            # Write header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for col_idx, column in enumerate(df_display.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = column.replace('_', ' ').title()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data
            for row_idx, row in enumerate(df_display.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        else:
            ws['A1'] = "No matched records found"
    
    def _create_exception_details_sheet(self, wb: Workbook, exception_records: List[Dict]):
        """Create detailed exception records worksheet"""
        ws = wb.create_sheet("Exception Details")
        
        if exception_records:
            df = pd.DataFrame(exception_records)
            
            # Select columns
            columns_to_show = [
                'transaction_date', 'store_name', 'exception_type', 'severity',
                'source_system', 'transaction_ref', 'amount', 'description',
                'recommended_action', 'status'
            ]
            
            columns_to_show = [col for col in columns_to_show if col in df.columns]
            df_display = df[columns_to_show]
            
            # Write header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            
            for col_idx, column in enumerate(df_display.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = column.replace('_', ' ').title()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data with conditional formatting
            for row_idx, row in enumerate(df_display.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Color code severity column
                    if df_display.columns[col_idx-1] == 'severity':
                        if value == 'CRITICAL':
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="C00000", bold=True)
                        elif value == 'HIGH':
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            cell.font = Font(color="FF6600")
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        else:
            ws['A1'] = "No exceptions found"
    
    def _create_store_summary_sheet(self, wb: Workbook, match_summary: Dict, exception_summary: Dict):
        """Create store-wise summary worksheet"""
        ws = wb.create_sheet("Store Summary")
        
        # This would aggregate data by store
        # For now, create placeholder
        ws['A1'] = "Store-wise Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Store Name"
        ws['B3'] = "Total Transactions"
        ws['C3'] = "Matched"
        ws['D3'] = "Exceptions"
        ws['E3'] = "Match Rate %"
        
        # Style header
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}3'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Placeholder data
        ws['A4'] = "Data will be aggregated by store in future version"
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def generate_html_report(self, 
                           match_summary: Dict, 
                           exception_summary: Dict) -> str:
        """
        Generate HTML report for email
        
        Args:
            match_summary: Summary from MatchingOrchestrator
            exception_summary: Summary from ExceptionHandler
        
        Returns:
            HTML content as string
        """
        logger.info("🌐 Generating HTML report...")
        
        # Get data
        batch_id = match_summary.get('batch_id', 'N/A')
        date_range = match_summary.get('date_range', {})
        source_counts = match_summary.get('source_counts', {})
        matches = match_summary.get('matches', {})
        financial = match_summary.get('financial', {})
        processing = match_summary.get('processing', {})
        
        total_exceptions = exception_summary.get('total_exceptions', 0)
        amount_at_risk = exception_summary.get('total_amount_at_risk', 0)
        by_type = exception_summary.get('by_type', {})
        by_severity = exception_summary.get('by_severity', {})
        
        # Determine overall status
        match_rate = financial.get('match_rate_percentage', 0)
        if match_rate >= 95 and total_exceptions == 0:
            status_color = '#28a745'
            status_icon = '✅'
            status_text = 'EXCELLENT'
        elif match_rate >= 90 and total_exceptions < 10:
            status_color = '#28a745'
            status_icon = '✅'
            status_text = 'GOOD'
        elif match_rate >= 80:
            status_color = '#ffc107'
            status_icon = '⚠️'
            status_text = 'NEEDS ATTENTION'
        else:
            status_color = '#dc3545'
            status_icon = '❌'
            status_text = 'CRITICAL'
        
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Matching Report - {batch_id}</title>
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f5f5f5;">
    <div style="max-width: 800px; margin: 0 auto; background-color: #ffffff;">
        
        <!-- Header -->
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 40px 30px; text-align: center;">
            <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 600;">
                Matching & Reconciliation Report
            </h1>
            <p style="color: #ffffff; margin: 10px 0 0 0; font-size: 14px; opacity: 0.9;">
                Batch: {batch_id}
            </p>
        </div>
        
        <!-- Status Banner -->
        <div style="background-color: {status_color}; padding: 20px; text-align: center;">
            <h2 style="color: #ffffff; margin: 0; font-size: 24px;">
                {status_icon} {status_text}
            </h2>
            <p style="color: #ffffff; margin: 5px 0 0 0; font-size: 14px;">
                Period: {date_range.get('start')} to {date_range.get('end')}
            </p>
        </div>
        
        <!-- Content -->
        <div style="padding: 30px;">
            
            <!-- Summary Cards Grid -->
            <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin-bottom: 30px;">
                
                <!-- Card: Match Rate -->
                <div style="background-color: #e8f5e9; border-radius: 8px; padding: 20px; text-align: center;">
                    <div style="font-size: 14px; color: #666; margin-bottom: 5px;">Match Rate</div>
                    <div style="font-size: 32px; font-weight: bold; color: #2e7d32;">{match_rate:.1f}%</div>
                </div>
                
                <!-- Card: Total Matched -->
                <div style="background-color: #e3f2fd; border-radius: 8px; padding: 20px; text-align: center;">
                    <div style="font-size: 14px; color: #666; margin-bottom: 5px;">Full Chain Matches</div>
                    <div style="font-size: 32px; font-weight: bold; color: #1976d2;">{matches.get('full_chain_matches', 0):,}</div>
                </div>
                
                <!-- Card: Exceptions -->
                <div style="background-color: #{'ffebee' if total_exceptions > 0 else 'f5f5f5'}; border-radius: 8px; padding: 20px; text-align: center;">
                    <div style="font-size: 14px; color: #666; margin-bottom: 5px;">Total Exceptions</div>
                    <div style="font-size: 32px; font-weight: bold; color: {'#c62828' if total_exceptions > 0 else '#666'};">{total_exceptions:,}</div>
                </div>
                
                <!-- Card: Amount at Risk -->
                <div style="background-color: #{'fff3e0' if amount_at_risk > 0 else 'f5f5f5'}; border-radius: 8px; padding: 20px; text-align: center;">
                    <div style="font-size: 14px; color: #666; margin-bottom: 5px;">Amount at Risk</div>
                    <div style="font-size: 24px; font-weight: bold; color: {'#f57c00' if amount_at_risk > 0 else '#666'};">₹{amount_at_risk:,.0f}</div>
                </div>
                
            </div>
            
            <!-- Source Data Summary -->
            <div style="margin-bottom: 30px;">
                <h3 style="color: #333; border-bottom: 2px solid #667eea; padding-bottom: 10px; margin-bottom: 15px;">
                    📊 Source Data Summary
                </h3>
                <table style="width: 100%; border-collapse: collapse;">
                    <tr style="background-color: #f5f5f5;">
                        <th style="padding: 10px; text-align: left; border-bottom: 1px solid #ddd;">Source</th>
                        <th style="padding: 10px; text-align: right; border-bottom: 1px solid #ddd;">Records</th>
                    </tr>
"""
        
        # Add source data rows
        for source, count in source_counts.items():
            html += f"""
                    <tr>
                        <td style="padding: 10px; border-bottom: 1px solid #eee;">{source.replace('_', ' ').title()}</td>
                        <td style="padding: 10px; text-align: right; border-bottom: 1px solid #eee; font-weight: 600;">{count:,}</td>
                    </tr>
"""
        
        html += """
                </table>
            </div>
            
            <!-- Match Results -->
            <div style="margin-bottom: 30px;">
                <h3 style="color: #333; border-bottom: 2px solid #667eea; padding-bottom: 10px; margin-bottom: 15px;">
                    🔗 Match Results
                </h3>
                <table style="width: 100%; border-collapse: collapse;">
                    <tr style="background-color: #f5f5f5;">
                        <th style="padding: 10px; text-align: left; border-bottom: 1px solid #ddd;">Match Type</th>
                        <th style="padding: 10px; text-align: right; border-bottom: 1px solid #ddd;">Count</th>
                    </tr>
"""
        
        # Add match result rows
        for match_type, count in matches.items():
            if count > 0:
                html += f"""
                    <tr>
                        <td style="padding: 10px; border-bottom: 1px solid #eee;">{match_type.replace('_', ' ').title()}</td>
                        <td style="padding: 10px; text-align: right; border-bottom: 1px solid #eee; font-weight: 600; color: #28a745;">{count:,}</td>
                    </tr>
"""
        
        # Exception Summary (if any)
        if total_exceptions > 0:
            html += """
            </table>
        </div>
        
        <!-- Exception Summary -->
        <div style="margin-bottom: 30px;">
            <h3 style="color: #c62828; border-bottom: 2px solid #ef5350; padding-bottom: 10px; margin-bottom: 15px;">
                ⚠️ Exception Summary
            </h3>
            <table style="width: 100%; border-collapse: collapse;">
                <tr style="background-color: #ffebee;">
                    <th style="padding: 10px; text-align: left; border-bottom: 1px solid #ddd;">Exception Type</th>
                    <th style="padding: 10px; text-align: right; border-bottom: 1px solid #ddd;">Count</th>
                </tr>
"""
            
            for exc_type, count in by_type.items():
                if count > 0:
                    html += f"""
                <tr>
                    <td style="padding: 10px; border-bottom: 1px solid #eee;">{exc_type.replace('_', ' ').title()}</td>
                    <td style="padding: 10px; text-align: right; border-bottom: 1px solid #eee; font-weight: 600; color: #c62828;">{count:,}</td>
                </tr>
"""
            
            html += """
            </table>
            
            <div style="margin-top: 15px; padding: 15px; background-color: #fff3e0; border-left: 4px solid #ff9800; border-radius: 4px;">
                <strong style="color: #e65100;">⚡ Action Required:</strong>
                <p style="margin: 5px 0 0 0; color: #666; font-size: 14px;">
                    Review exception details in the attached Excel report. Critical and High severity items require immediate attention.
                </p>
            </div>
        </div>
"""
        else:
            html += """
            </table>
        </div>
        
        <div style="padding: 20px; background-color: #e8f5e9; border-radius: 8px; text-align: center; margin-bottom: 30px;">
            <div style="font-size: 48px; margin-bottom: 10px;">🎉</div>
            <div style="font-size: 18px; color: #2e7d32; font-weight: 600;">
                Perfect Match! No Exceptions Found
            </div>
        </div>
"""
        
        html += f"""
            <!-- Processing Info -->
            <div style="background-color: #f5f5f5; padding: 15px; border-radius: 8px; font-size: 13px; color: #666;">
                <div style="margin-bottom: 5px;">
                    <strong>Records Processed:</strong> {processing.get('total_records_processed', 0):,}
                </div>
                <div style="margin-bottom: 5px;">
                    <strong>Processing Time:</strong> {processing.get('processing_time_seconds', 0):.2f}s
                </div>
                <div>
                    <strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                </div>
            </div>
            
        </div>
        
        <!-- Footer -->
        <div style="background-color: #333; padding: 20px; text-align: center;">
            <p style="color: #ffffff; margin: 0; font-size: 12px;">
                Automated Matching & Reconciliation System
            </p>
            <p style="color: #999; margin: 5px 0 0 0; font-size: 11px;">
                Devyani Data Team - Smart Uploader
            </p>
        </div>
        
    </div>
</body>
</html>
"""
        
        logger.info("  ✅ HTML report generated")
        return html
    
    def save_html_report(self, html_content: str, batch_id: str) -> str:
        """
        Save HTML report to file
        
        Args:
            html_content: HTML string
            batch_id: Batch ID for filename
        
        Returns:
            Path to saved HTML file
        """
        self.ensure_reports_directory()
        
        filename = f"matching_report_{batch_id}.html"
        filepath = os.path.join(self.reports_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"  ✅ HTML report saved: {filepath}")
        return filepath


if __name__ == '__main__':
    # Quick test
    generator = ReportGenerator()
    
    # Sample data
    match_summary = {
        'batch_id': 'MATCH_20251112_TEST',
        'date_range': {'start': '2024-12-01', 'end': '2024-12-07'},
        'source_counts': {
            'pos_records': 1500,
            'trm_records': 1400,
            'mpr_upi_records': 800,
            'mpr_card_records': 600,
            'bank_records': 250
        },
        'matches': {
            'trm_mpr_upi': 780,
            'trm_mpr_card': 590,
            'full_chain_matches': 1350
        },
        'financial': {
            'total_pos_amount': 5000000.00,
            'match_rate_percentage': 92.5
        },
        'processing': {
            'total_records_processed': 4550,
            'processing_time_seconds': 45.3
        }
    }
    
    exception_summary = {
        'total_exceptions': 25,
        'total_amount_at_risk': 125000.00,
        'by_type': {
            'MISSING_TRM': 10,
            'MISSING_MPR': 8,
            'MISSING_BANK': 5,
            'AMOUNT_MISMATCH': 2
        },
        'by_severity': {
            'CRITICAL': 3,
            'HIGH': 7,
            'MEDIUM': 10,
            'LOW': 5
        }
    }
    
    # Generate HTML
    html = generator.generate_html_report(match_summary, exception_summary)
    html_path = generator.save_html_report(html, 'MATCH_20251112_TEST')
    
    print(f"\n✅ HTML report generated: {html_path}")
    print(f"📊 Open in browser to view")
