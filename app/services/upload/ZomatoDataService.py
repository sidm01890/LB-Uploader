"""
Zomato Data Service - OPTIMIZED
Handles Zomato data file uploads with performance optimizations:
1. Increased batch size (5000)
2. Single DB connection for entire upload
3. executemany() instead of row-by-row execute
4. Pre-computed column mappings
5. Sets for O(1) type checking
"""
import logging
import os
import traceback
import time
import random
from datetime import datetime, timedelta
from typing import List, Dict, Set, Optional, Tuple

import pandas as pd
import numpy as np
try:
    from mysql.connector import InterfaceError  # type: ignore
except ImportError:  # pragma: no cover
    InterfaceError = Exception
from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session

from app.core.config import config
from app.core.enum.DataSource import DataSource as DS
from app.services.DataSourceAbstractService import DataSourceService
from app.core.database import db_manager
from app.services.mongodb_service import save_uploaded_sheet

log = logging.getLogger(__name__)


# ============================================================================
# OPTIMIZATION: Define column type sets at module level for O(1) lookup
# ============================================================================
DATE_COLUMNS: Set[str] = {'order_date', 'payout_date', 'utr_date'}

NUMERIC_COLUMNS: Set[str] = {
    'amount', 'net_amount', 'final_amount', 'customer_compensation',
    'pro_discount', 'commission_rate', 'zvd', 'tcs_amount',
    'rejection_penalty_charge', 'user_credits_charge', 'promo_recovery_adj',
    'icecream_handling', 'icecream_deductions', 'order_support_cost',
    'pro_discount_passthrough', 'mvd', 'delivery_charge', 'bill_subtotal',
    'commission_value', 'tax_rate', 'taxes_zomato_fee', 'credit_note_amount',
    'tds_amount', 'pgcharge', 'logistics_charge', 'mdiscount',
    'customer_discount', 'cancellation_refund', 'total_amount',
    'total_voucher', 'source_tax', 'tax_paid_by_customer',
    'charged_by_zomato_agt_rejected_orders', 'percent', 'pg_chgs_percent',
    'merchant_delivery_charge', 'merchant_pack_charge'
}

# String columns that should remain as strings (VARCHAR in database)
# Note: pg_applied_on, order_id, res_id, city_id are VARCHAR in DB, not numeric/int
STRING_COLUMNS: Set[str] = {
    'pg_applied_on', 'order_id', 'res_id', 'city_id', 'ls',
    'uid', 'store_code', 'sap_code', 'store_name', 'city_name', 
    'res_name', 'status', 'utr_number', 'service_id', 'action',
    'payment_method', 'promo_code'
}


class ZomatoDataService(DataSourceService):
    """Service for uploading Zomato data files - OPTIMIZED VERSION"""
    
    # OPTIMIZATION: Increased batch size from 2000 to 5000
    CHUNK_SIZE = 5000
    data_upload_dir = config.data_upload_dir
    directory_path = os.path.join(data_upload_dir, "three_po", "zomato")

    @staticmethod
    def _normalize_column_name(column_name: str) -> str:
        """Normalize column name for comparison"""
        if not isinstance(column_name, str):
            return str(column_name)
        normalized = column_name.replace('\n', ' ').replace('\r', ' ')
        normalized = ''.join(c.lower() for c in normalized if c.isalnum() or c.isspace() or c == '%')
        normalized = ' '.join(normalized.strip().split())
        return normalized

    INVALID_NUMERIC_TOKENS = {
        '', 'na', '#n/a', 'n/a', '--', '-', 'null', 'nil', 'nan', 'none'
    }

    def _coerce_to_float(self, value):
        """Safely convert incoming value to float, returning None when not possible."""
        if value is None:
            return None
        
        # Handle pandas/numpy NaN values
        if pd.isna(value):
            return None

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            normalized = cleaned.lower()
            if normalized in self.INVALID_NUMERIC_TOKENS:
                return None

            # Remove common thousands separators
            cleaned = cleaned.replace(',', '')

            # Handle percentage strings like "12.5%"
            if cleaned.endswith('%'):
                number_part = cleaned[:-1]
                if not number_part:
                    return None
                try:
                    return float(number_part)
                except ValueError:
                    return None

            try:
                return float(cleaned)
            except ValueError:
                return None

        # Handle Decimal types from openpyxl/pandas
        if hasattr(value, '__float__'):
            try:
                return float(value)
            except (TypeError, ValueError, OverflowError):
                return None

        try:
            return float(value)
        except (TypeError, ValueError, OverflowError):
            return None

    def _coerce_to_int(self, value):
        """Safely convert incoming value to int, returning None when not possible."""
        if value is None:
            return None
        
        if pd.isna(value):
            return None

        if isinstance(value, str):
            cleaned = value.strip()
            if not cleaned:
                return None
            normalized = cleaned.lower()
            if normalized in self.INVALID_NUMERIC_TOKENS:
                return None

            # Remove common thousands separators
            cleaned = cleaned.replace(',', '')

            try:
                # Try float first, then convert to int (handles "123.0" -> 123)
                float_val = float(cleaned)
                return int(float_val)
            except ValueError:
                return None

        try:
            if isinstance(value, float):
                # Handle NaN
                if pd.isna(value):
                    return None
                return int(value)
            return int(value)
        except (TypeError, ValueError, OverflowError):
            return None

    def _convert_to_datetime(self, value):
        """Convert value to datetime if possible."""
        if pd.isna(value) or value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                excel_date = float(value)
                if excel_date > 59:
                    excel_date -= 1
                return datetime(1899, 12, 30) + timedelta(days=excel_date)
            elif isinstance(value, str):
                for fmt in ['%d.%m.%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']:
                    try:
                        return pd.to_datetime(value, format=fmt)
                    except:
                        continue
                return pd.to_datetime(value)
            elif isinstance(value, datetime):
                return value
            return None
        except:
            return None
    
    def _extract_utr_from_text(self, text_value):
        """Extract UTR number from text field"""
        if pd.isna(text_value) or not isinstance(text_value, str):
            return None
        import re
        pattern = r'CITI[N]?(\d+)'
        match = re.search(pattern, text_value, re.IGNORECASE)
        if match:
            return match.group(0)
            return None

    async def upload_file(self, files: List[UploadFile]) -> List[str]:
        """Save uploaded files to disk using streaming (handles large files efficiently)"""
        os.makedirs(self.directory_path, exist_ok=True)
        uploaded_file_paths = []
        
        for file in files:
            filename = file.filename
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                content_type = getattr(file, 'content_type', '')
                if 'excel' in content_type or 'spreadsheet' in content_type:
                    ext = '.xlsx'
                elif 'csv' in content_type:
                    ext = '.csv'
                else:
                    ext = '.xlsx'
                filename = f"upload_{timestamp}{ext}"
            
            file_path = os.path.join(self.directory_path, filename)
            chunk_size = 1024 * 1024  # 1MB chunks
            total_size = 0
            
            # Stream write using async read (prevents loading entire file into memory)
            with open(file_path, "wb") as f:
                while True:
                    chunk = await file.read(chunk_size)
                    if not chunk:
                        break
                    f.write(chunk)
                    total_size += len(chunk)
            
            # Reset file pointer for potential future reads
            await file.seek(0)
            
            absolute_path = os.path.abspath(file_path)
            log.info(f"File saved at path: {absolute_path} (size: {total_size} bytes)")
            uploaded_file_paths.append(absolute_path)
        
        return uploaded_file_paths

    def _build_normalized_mapping(self, fieldsMap: Dict[str, List[str]], 
                                   mapped_excel_columns: List[str]) -> Dict[str, str]:
        """
        OPTIMIZATION: Pre-compute normalized column name mapping
        Returns: {normalized_excel_col: original_excel_col}
        """
        normalized_to_original = {}
        for excel_col in mapped_excel_columns:
            normalized = self._normalize_column_name(excel_col)
            normalized_to_original[normalized] = excel_col
        return normalized_to_original

    def _process_row(self, row_dict: Dict, fieldsMap: Dict[str, List[str]], 
                     normalized_mapping: Dict[str, str],
                     header_normalized: Dict[str, str],
                     credit_col: Optional[str], debit_col: Optional[str], 
                     local_crcy_col: Optional[str]) -> Optional[Dict]:
        """Process a single row and return data dict or None if invalid"""
        data = {}
        raw_values = {}
        
        for excel_col, value in row_dict.items():
            raw_values[excel_col] = value
        
        # Map columns using pre-computed normalized mapping
        for excel_col, value in row_dict.items():
            normalized_excel_col = header_normalized.get(excel_col)
            if normalized_excel_col is None:
                normalized_excel_col = self._normalize_column_name(str(excel_col))
            
            matching_col = normalized_mapping.get(normalized_excel_col)
            
            if matching_col:
                db_cols = fieldsMap[matching_col]
                if value is None:
                    for db_col in db_cols:
                        data[db_col] = None
                else:
                    for db_col in db_cols:
                        # OPTIMIZATION: Use set lookup O(1) instead of list O(n)
                        if db_col in DATE_COLUMNS:
                            converted = None
                            if isinstance(value, (int, float)) and value and not pd.isna(value):
                                int_val = int(value)
                                if 19000101 <= int_val <= 21001231:
                                    try:
                                        converted = datetime.strptime(str(int_val), "%Y%m%d")
                                    except Exception:
                                        converted = None
                            if converted is None:
                                converted = self._convert_to_datetime(value)
                            data[db_col] = converted.date() if converted else None
                        elif db_col == 'utr_number' and str(excel_col) == 'Text':
                            data[db_col] = self._extract_utr_from_text(value)
                        elif db_col in NUMERIC_COLUMNS:
                            # Convert to float - ensure it's actually a float, not a string representation
                            converted = self._coerce_to_float(value)
                            # If conversion failed but value exists, try direct conversion
                            if converted is None and value is not None and not pd.isna(value):
                                try:
                                    # Last resort: try direct float conversion
                                    converted = float(value)
                                except (ValueError, TypeError):
                                    converted = None
                            # CRITICAL: Ensure it's actually a float type, not a string that looks like a number
                            if converted is not None:
                                # Double-check it's not a string
                                if isinstance(converted, str):
                                    try:
                                        converted = float(converted)
                                    except (ValueError, TypeError):
                                        converted = None
                            data[db_col] = converted
                        elif db_col in STRING_COLUMNS:
                            # String columns (VARCHAR in database) - MUST be string, never int/float
                            if value is None or pd.isna(value):
                                data[db_col] = None
                            else:
                                # CRITICAL: Convert to string - handle both numeric and string inputs
                                # Excel may read these as integers, but DB expects VARCHAR
                                if isinstance(value, (int, float)):
                                    # Convert numeric to string
                                    data[db_col] = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
                                else:
                                    data[db_col] = str(value) if value is not None else None
                        else:
                            # For other string columns, ensure it's a string or None
                            if value is None or pd.isna(value):
                                data[db_col] = None
                            else:
                                data[db_col] = str(value) if not isinstance(value, str) else value
        
        # Handle amount calculations
        credit_val = raw_values.get(credit_col) if credit_col else None
        debit_val = raw_values.get(debit_col) if debit_col else None
        local_crcy_val = raw_values.get(local_crcy_col) if local_crcy_col else None
        
        try:
            credit = float(credit_val) if credit_val is not None else 0
            debit = float(debit_val) if debit_val is not None else 0
            local_crcy = float(local_crcy_val) if local_crcy_val is not None else 0
        except (ValueError, TypeError):
            credit = debit = local_crcy = 0
        
        calculated_amount = None
        if credit_col or debit_col or local_crcy_col:
            if credit > 0:
                calculated_amount = credit
            elif local_crcy < 0:
                calculated_amount = abs(local_crcy)
            elif local_crcy != 0:
                calculated_amount = local_crcy
            elif debit > 0:
                calculated_amount = abs(debit)
            else:
                calculated_amount = 0
        
        if calculated_amount is not None:
            data['amount'] = calculated_amount
            data['net_amount'] = calculated_amount
            data['final_amount'] = calculated_amount
        else:
            if 'amount' not in data:
                data['amount'] = None
            if 'net_amount' not in data:
                data['net_amount'] = None
            if 'final_amount' not in data:
                data['final_amount'] = None
        
        if 'status' not in data or data['status'] is None:
            data['status'] = 'completed' if credit > 0 else 'pending'
        
        # Generate UID - use order_id if available, otherwise use fallback
        order_id = data.get('order_id')
        if order_id:
            data['uid'] = str(order_id)
        else:
            # Fallback UID generation for rows without order_id
            # Try res_id first, then use a combination of other fields
            res_id = data.get('res_id')
            order_date = data.get('order_date')
            
            if res_id and order_date:
                # Use res_id + order_date as fallback UID
                try:
                    if isinstance(order_date, datetime):
                        date_str = order_date.strftime('%Y%m%d')
                    elif isinstance(order_date, str):
                        date_str = order_date.replace('-', '').replace('/', '')[:8]
                    else:
                        date_str = str(order_date)[:8]
                    data['uid'] = f"{res_id}_{date_str}"
                except Exception:
                    data['uid'] = f"ZOMATO_{res_id}_{int(time.time())}"
            elif res_id:
                # Use res_id with timestamp to ensure uniqueness
                data['uid'] = f"ZOMATO_{res_id}_{int(time.time() * 1000)}"
            else:
                # Last resort: use timestamp + random component
                data['uid'] = f"ZOMATO_{int(time.time() * 1000000)}_{random.randint(1000, 9999)}"
        
        return data

    def _bulk_insert(self, cursor, batch: List[Dict], table_name: str = 'zomato') -> int:
        """
        OPTIMIZATION: Bulk insert using executemany with prepared statement
        """
        if not batch:
            return 0
        
        # Get columns from first record (all records should have same columns)
        columns = list(batch[0].keys())
        placeholders = ', '.join(['%s'] * len(columns))
        update_clause = ', '.join([f"{col} = VALUES({col})" for col in columns if col != 'uid'])
        
        query = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders}) ON DUPLICATE KEY UPDATE {update_clause}"
        
        # Prepare all values as list of tuples with final sanitization
        values_list = []
        for record in batch:
            sanitized_record = {}
            for col in columns:
                value = record.get(col)
                
                # Final sanitization: ensure proper types for MySQL
                # Check for None or NaN values first
                if value is None:
                    sanitized_record[col] = None
                elif isinstance(value, float) and (pd.isna(value) or np.isnan(value)):
                    sanitized_record[col] = None
                elif col in NUMERIC_COLUMNS:
                    # Ensure numeric columns are float or None, never strings
                    # This is critical - MySQL connector fails if we pass strings to numeric columns
                    if isinstance(value, str):
                        converted = self._coerce_to_float(value)
                        if converted is None:
                            # If conversion failed, log and set to None
                            log.warning(f"Failed to convert {col}='{value}' to float, setting to None")
                            sanitized_record[col] = None
                        else:
                            sanitized_record[col] = converted
                    elif isinstance(value, (int, float)):
                        # Check for NaN
                        if isinstance(value, float) and (pd.isna(value) or np.isnan(value)):
                            sanitized_record[col] = None
                        else:
                            sanitized_record[col] = float(value)
                    elif hasattr(value, '__float__'):
                        # Handle Decimal and other numeric types from openpyxl
                        try:
                            sanitized_record[col] = float(value)
                        except (TypeError, ValueError, OverflowError):
                            sanitized_record[col] = None
                    else:
                        # Unknown type - try to convert, otherwise None
                        try:
                            sanitized_record[col] = float(value) if value is not None else None
                        except (TypeError, ValueError):
                            sanitized_record[col] = None
                elif col in STRING_COLUMNS:
                    # String columns (VARCHAR in database) - MUST be string, never int/float
                    # CRITICAL: MySQL connector is very strict - ALL values must be strings or None
                    if value is None or (isinstance(value, float) and (pd.isna(value) or np.isnan(value))):
                        sanitized_record[col] = None
                    else:
                        # CRITICAL: Convert to string - handle both numeric and string inputs
                        # Excel may read these as integers, but DB expects VARCHAR
                        # We MUST convert ALL numeric types to strings here
                        if isinstance(value, int):
                            sanitized_record[col] = str(value)
                        elif isinstance(value, float):
                            # Convert float to string, handling integer floats
                            if value.is_integer():
                                sanitized_record[col] = str(int(value))
                            else:
                                sanitized_record[col] = str(value)
                        else:
                            # Already a string or other type - convert to string
                            sanitized_record[col] = str(value) if value is not None else None
                elif col in DATE_COLUMNS:
                    # Date columns should already be date objects or None
                    sanitized_record[col] = value
                else:
                    # String columns: ensure it's a string or None
                    if value is None or (isinstance(value, float) and (pd.isna(value) or np.isnan(value))):
                        sanitized_record[col] = None
                    elif isinstance(value, (int, float)):
                        # Convert numbers to strings for string columns
                        if isinstance(value, float) and (pd.isna(value) or np.isnan(value)):
                            sanitized_record[col] = None
                        else:
                            sanitized_record[col] = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
                    else:
                        sanitized_record[col] = str(value) if value is not None else None
                        
            # CRITICAL: Final type check before creating tuple - ensure STRING_COLUMNS are strings
            for col in columns:
                if col in STRING_COLUMNS:
                    val = sanitized_record.get(col)
                    if val is not None and not isinstance(val, str):
                        # Last chance conversion - this should not happen if sanitization worked
                        if isinstance(val, int):
                            sanitized_record[col] = str(val)
                        elif isinstance(val, float):
                            sanitized_record[col] = str(int(val)) if val.is_integer() else str(val)
                        else:
                            sanitized_record[col] = str(val)
                        log.warning(f"⚠️ Last-chance conversion: {col} from {type(val).__name__} to str")
            
            values = tuple(sanitized_record.get(col) for col in columns)
            values_list.append(values)
        
        # OPTIMIZATION: Use executemany instead of execute in loop
        # Final validation pass - ensure proper types match database schema
        log.info(f"🔍 Final validation pass: checking {len(values_list)} rows, {len(columns)} columns")
        type_issues = []
        for idx, row_tuple in enumerate(values_list):
            row_dict = dict(zip(columns, row_tuple))
            for col in columns:
                val = row_dict[col]
                if col in NUMERIC_COLUMNS:
                    # Numeric columns must be float or None, never strings
                    if isinstance(val, str):
                        # Last chance conversion
                        converted = self._coerce_to_float(val)
                        if converted is not None:
                            # Update the tuple value
                            col_idx = columns.index(col)
                            row_list = list(row_tuple)
                            row_list[col_idx] = converted
                            values_list[idx] = tuple(row_list)
                            log.warning(f"⚠️ Row {idx}, col {col}: Converted string '{val}' to float {converted} at last moment")
                            type_issues.append(f"Row {idx}, {col}: str->float")
                        else:
                            log.error(f"❌ Row {idx}, col {col}: Cannot convert string '{val}' to float")
                            type_issues.append(f"Row {idx}, {col}: FAILED str->float")
                    elif val is not None and not isinstance(val, (int, float)):
                        # Try to convert unknown types
                        try:
                            converted = float(val)
                            col_idx = columns.index(col)
                            row_list = list(row_tuple)
                            row_list[col_idx] = converted
                            values_list[idx] = tuple(row_list)
                            log.warning(f"⚠️ Row {idx}, col {col}: Converted {type(val).__name__} '{val}' to float {converted}")
                            type_issues.append(f"Row {idx}, {col}: {type(val).__name__}->float")
                        except (TypeError, ValueError):
                            log.error(f"❌ Row {idx}, col {col}: Cannot convert {type(val).__name__} '{val}' to float")
                            type_issues.append(f"Row {idx}, {col}: FAILED {type(val).__name__}->float")
                elif col in STRING_COLUMNS:
                    # String columns must be string or None, never numeric types
                    # CRITICAL: This should never happen if sanitization worked, but double-check
                    if val is not None and not isinstance(val, str):
                        # Convert numeric/other types to string
                        col_idx = columns.index(col)
                        row_list = list(row_tuple)
                        # Handle both int and float - convert to string
                        if isinstance(val, int):
                            converted_val = str(val)
                        elif isinstance(val, float):
                            converted_val = str(int(val)) if val.is_integer() else str(val)
                        else:
                            converted_val = str(val) if val is not None else None
                        row_list[col_idx] = converted_val
                        values_list[idx] = tuple(row_list)
                        log.error(f"❌ Row {idx}, col {col}: CRITICAL - Found {type(val).__name__} '{val}' in STRING_COLUMN, converted to '{converted_val}'")
                        type_issues.append(f"Row {idx}, {col}: {type(val).__name__}->str")
        
        if type_issues:
            log.info(f"🔍 Type conversions made: {len(type_issues)} issues fixed")
        else:
            log.info(f"✅ No type conversion issues found - all types match schema")
        
        # Log first row types for debugging
        if values_list:
            first_row = dict(zip(columns, values_list[0]))
            log.info(f"🔍 First row final types: {[(col, type(val).__name__) for col, val in first_row.items() if col in NUMERIC_COLUMNS or col in STRING_COLUMNS]}")
        
        # Use execute in loop instead of executemany for more lenient type handling
        # MySQL connector's executemany is very strict about types and can fail even with correct conversions
        try:
            rows_inserted = 0
            for row_values in values_list:
                try:
                    cursor.execute(query, row_values)
                    rows_inserted += 1
                except InterfaceError as row_error:  # type: ignore[arg-type]
                    # Log the problematic row for debugging
                    row_dict = dict(zip(columns, row_values))
                    log.error(f"❌ Failed to insert row {rows_inserted + 1}: {str(row_error)}")
                    log.error(f"   Problematic row data: {row_dict}")
                    # Check for type issues in this specific row
                    for col, val in zip(columns, row_values):
                        if col in NUMERIC_COLUMNS and isinstance(val, str):
                            log.error(f"   ⚠️ Column {col} is string but should be float: {repr(val)}")
                        elif col in STRING_COLUMNS and not isinstance(val, (str, type(None))):
                            log.error(f"   ⚠️ Column {col} is {type(val).__name__} but should be str: {repr(val)}")
                    raise  # Re-raise to trigger outer exception handler
            # Note: Commit is handled by the calling code at the connection level
            log.info(f"✅ Successfully inserted {rows_inserted} rows using execute loop")
            return rows_inserted
        except InterfaceError as e:  # type: ignore[arg-type]
            sample = values_list[0] if values_list else ()
            sample_dict = dict(zip(columns, sample))
            
            # Check for type mismatches
            invalid_numeric = {
                col: (type(val).__name__, repr(val))
                for col, val in sample_dict.items()
                if col in NUMERIC_COLUMNS and isinstance(val, str)
            }
            invalid_string = {
                col: (type(val).__name__, repr(val))
                for col, val in sample_dict.items()
                if col in STRING_COLUMNS and not isinstance(val, (str, type(None)))
            }
            
            log.error(f"Failed batch sample (first row): {sample_dict}")
            log.error(f"Column type snapshot: {[ (col, type(val).__name__) for col, val in sample_dict.items() ]}")
            
            if invalid_numeric:
                log.error(f"❌ Numeric columns still strings: {invalid_numeric}")
            if invalid_string:
                log.error(f"❌ String columns with wrong type: {invalid_string}")
            
            # Check all rows for type issues
            if not invalid_numeric and not invalid_string:
                for idx, row in enumerate(values_list):
                    row_dict = dict(zip(columns, row))
                    bad_numeric = {
                        col: (type(val).__name__, repr(val))
                        for col, val in row_dict.items()
                        if col in NUMERIC_COLUMNS and isinstance(val, str)
                    }
                    bad_string = {
                        col: (type(val).__name__, repr(val))
                        for col, val in row_dict.items()
                        if col in STRING_COLUMNS and not isinstance(val, (str, type(None)))
                    }
                    if bad_numeric or bad_string:
                        log.error(f"❌ Offending row {idx}: numeric={bad_numeric}, string={bad_string}")
                        break
            raise

    def upload(self, files: List[str], db: Session, username: str) -> int:
        """
        Upload and process Zomato data files - OPTIMIZED VERSION
        
        Optimizations:
        1. Single DB connection for entire upload
        2. Pre-computed column mappings
        3. executemany() for bulk inserts
        4. Set-based type checking
        """
        total_rows_inserted = 0
        
        try:
            # Get table column mappings from database
            with db_manager.get_mysql_connector() as connection:
                cursor = connection.cursor(dictionary=True)
                
                cursor.execute("""
                    SELECT excel_column_name, db_column_name 
                    FROM table_columns_mapping 
                    WHERE data_source = %s
                """, (DS.ZOMATO.name,))
                mappings = cursor.fetchall()
                
                if not mappings:
                    raise HTTPException(
                        status_code=400,
                        detail=f"No field mappings found for {DS.ZOMATO.name}"
                    )
                
                # Build mapping
                fieldsMap = {}
                for m in mappings:
                    excel_col = m['excel_column_name'].strip()
                    db_col = m['db_column_name'].strip()
                    if excel_col not in fieldsMap:
                        fieldsMap[excel_col] = []
                    fieldsMap[excel_col].append(db_col)
                
                mapped_excel_columns = list(fieldsMap.keys())
                
                cursor.execute("""
                    SELECT uid_columns FROM data_source WHERE name = %s
                """, (DS.ZOMATO.name,))
                data_source = cursor.fetchone()
                
                if not data_source or not data_source.get('uid_columns'):
                    raise HTTPException(
                        status_code=400,
                        detail=f"No uid_columns found in DataSource table for {DS.ZOMATO.name}"
                    )
                
                uid_columns = [col.strip() for col in data_source['uid_columns'].split(',')]
                log.info(f"Using uid columns: {uid_columns}")
            
            # OPTIMIZATION: Pre-compute normalized mapping ONCE
            normalized_mapping = self._build_normalized_mapping(fieldsMap, mapped_excel_columns)
            log.info(f"Pre-computed {len(normalized_mapping)} column mappings")
            
            # Process each file
            for file_path in files:
                log.info(f"Processing file: {file_path}")
                
                import openpyxl
                
                try:
                    if file_path.endswith(('.xlsx', '.xls')):
                        workbook = openpyxl.load_workbook(file_path, read_only=True, keep_links=False, data_only=True)
                        sheet_names = workbook.sheetnames
                        
                        if len(sheet_names) > 1:
                            # Priority 1: Exact match "Store wise details" (case-insensitive)
                            target_sheets = [s for s in sheet_names if s.strip().lower() == "store wise details"]
                            
                            # Priority 2: If no exact match, look for sheets starting with "store wise details"
                            # but prefer the one WITHOUT a suffix (e.g., prefer "Store wise details" over "Store wise details-2")
                            if not target_sheets:
                                matching_sheets = [s for s in sheet_names if s.strip().lower().startswith("store wise details")]
                                if matching_sheets:
                                    # Sort to get the shortest name first (likely the main one without suffix)
                                    matching_sheets.sort(key=len)
                                    target_sheets = [matching_sheets[0]]  # Only take the first/main one
                                    log.info(f"Found multiple 'Store wise details' sheets, using: {target_sheets[0]}")
                            
                            if not target_sheets:
                                log.warning(f"File {file_path} has multiple sheets but none match 'Store wise details'; skipping.")
                                workbook.close()
                                continue
                        else:
                            target_sheets = sheet_names
                        
                        # Collect raw data for MongoDB
                        all_raw_data = []
                        file_headers = []
                        
                        # TEMPORARILY DISABLED: MySQL connection for MongoDB testing only
                        # # OPTIMIZATION: Single connection for ALL sheets and batches
                        # with db_manager.get_mysql_connector() as db_connection:
                        #     db_cursor = db_connection.cursor()
                        
                        for sheet_name in target_sheets:
                                log.info(f"Reading sheet: {sheet_name}")
                                
                                worksheet = workbook[sheet_name]
                                rows = worksheet.iter_rows(values_only=True)
                                header = next(rows)
                                
                                # Store headers for MongoDB (first sheet only)
                                if not file_headers:
                                    file_headers = [str(col) if col is not None else '' for col in header]
                                
                                col_map = {i: col for i, col in enumerate(header) if col is not None}
                                
                                # OPTIMIZATION: Pre-compute header normalization
                                header_normalized = {col: self._normalize_column_name(str(col)) for col in col_map.values()}
                                
                                # OPTIMIZATION: Find special columns ONCE
                                credit_col = None
                                debit_col = None
                                local_crcy_col = None
                                for col in col_map.values():
                                    normalized = self._normalize_column_name(str(col))
                                    if normalized == 'credit':
                                        credit_col = col
                                    elif normalized == 'debit':
                                        debit_col = col
                                    elif normalized == 'local crcy amt':
                                        local_crcy_col = col
                                
                                batch_to_insert = []
                                processed_count = 0
                                batch_count = 0
                                skipped_count = 0
                                
                                for row in rows:
                                    row_dict = {col_map[i]: val for i, val in enumerate(row) if i in col_map}
                                    
                                    # Collect raw data for MongoDB (before processing)
                                    all_raw_data.append(row_dict.copy())
                                    
                                    # TEMPORARILY DISABLED: MySQL insertion for MongoDB testing only
                                    # data = self._process_row(
                                    #     row_dict, fieldsMap, normalized_mapping, header_normalized,
                                    #     credit_col, debit_col, local_crcy_col
                                    # )
                                    
                                    # if data:
                                    #     batch_to_insert.append(data)
                                    # else:
                                    #     skipped_count += 1
                                    #     if skipped_count <= 5:  # Log first 5 skipped rows for debugging
                                    #         log.warning(f"Skipped row (no data returned): {row_dict.get('order_id', 'N/A')}")
                                    
                                    processed_count += 1
                                    
                                    # Log progress every 100k rows
                                    if processed_count % 100000 == 0:
                                        log.info(f"Processed {processed_count:,} rows for MongoDB, {skipped_count:,} skipped")
                                    
                                    # TEMPORARILY DISABLED: MySQL batch insertion
                                    # # Insert in batches
                                    # if len(batch_to_insert) >= self.CHUNK_SIZE:
                                    #     batch_count += 1
                                    #     log.info(f"Inserting batch {batch_count}: {len(batch_to_insert)} records")
                                    #     
                                    #     rows_inserted = self._bulk_insert(db_cursor, batch_to_insert)
                                    #     db_connection.commit()
                                    #     total_rows_inserted += rows_inserted
                                    #     
                                    #     log.info(f"Batch {batch_count} complete. Total: {total_rows_inserted:,} rows")
                                    #     batch_to_insert = []
                                
                                # TEMPORARILY DISABLED: MySQL final batch insertion
                                # # Insert remaining records
                                # if batch_to_insert:
                                #     batch_count += 1
                                #     log.info(f"Inserting final batch {batch_count}: {len(batch_to_insert)} records")
                                #     
                                #     rows_inserted = self._bulk_insert(db_cursor, batch_to_insert)
                                #     db_connection.commit()
                                #     total_rows_inserted += rows_inserted
                                #     
                                #     log.info(f"Final batch complete. Total: {total_rows_inserted:,} rows")
                                
                                log.info(f"Sheet {sheet_name}: Processed {processed_count:,} rows for MongoDB")
                        
                        workbook.close()
                        
                        # Save raw data to MongoDB
                        if all_raw_data:
                            try:
                                filename = os.path.basename(file_path)
                                file_size = os.path.getsize(file_path)
                                file_type = filename.split('.')[-1] if '.' in filename else 'xlsx'
                                
                                # Transform row-based data to column-based data
                                # From: [{"col1": "val1", "col2": "val2"}, {"col1": "val3", "col2": "val4"}]
                                # To: {"col1": ["val1", "val3"], "col2": ["val2", "val4"]}
                                column_based_data = {}
                                if all_raw_data:
                                    # Get all column names from the first row
                                    all_columns = set()
                                    for row in all_raw_data:
                                        all_columns.update(row.keys())
                                    
                                    # Initialize columns with empty lists
                                    for col in all_columns:
                                        column_based_data[col] = []
                                    
                                    # Fill column arrays with values from each row
                                    for row in all_raw_data:
                                        for col in all_columns:
                                            column_based_data[col].append(row.get(col))
                                
                                save_uploaded_sheet(
                                    filename=filename,
                                    datasource=DS.ZOMATO.name,
                                    table_name="zomato",
                                    file_size=file_size,
                                    file_type=file_type,
                                    headers=file_headers,
                                    raw_data=column_based_data,
                                    uploaded_by=username,
                                    column_mapping=None
                                )
                                log.info(f"✅ Saved {len(all_raw_data)} rows of raw data to MongoDB (column-based format)")
                            except Exception as mongo_error:
                                log.warning(f"⚠️ Failed to save raw data to MongoDB (continuing): {mongo_error}")

                    elif file_path.endswith('.csv'):
                        log.warning("CSV processing not yet optimized")
                        pass
                         
                    else:
                        raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_path}")
                
                except Exception as e:
                    raise HTTPException(status_code=400, detail=f"Error reading file {file_path}: {str(e)}")
            
            # TEMPORARILY DISABLED: Return 0 for MySQL rows since we're only testing MongoDB
            log.info(f"Zomato upload completed. MongoDB: {len(all_raw_data) if 'all_raw_data' in locals() else 0} rows saved")
            return 0  # Return 0 since MySQL insertion is disabled for MongoDB testing
            
        except HTTPException as he:
            log.error(f"HTTP Error during data upload: {he.detail}")
            raise he
        except Exception as e:
            log.error(f"Error during data upload: {e}")
            traceback.print_exc()
            # TEMPORARILY DISABLED: No database operations for MongoDB testing
            # db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"Internal server error during data upload: {str(e)}"
            )
        # TEMPORARILY DISABLED: No database connection to close
        # finally:
        #     db.close()
